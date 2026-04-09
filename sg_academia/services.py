import pandas as pd
from datetime import date, timedelta
from pathlib import Path
from database import get_connection, DURACOES, VALORES

REPORTS_DIR = Path(__file__).resolve().parent / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


# ── Status automático ────────────────────────────────────────────
def atualizar_status():
    hoje = date.today().isoformat()
    conn = get_connection()
    # Atualiza status_plano pelo vencimento
    conn.execute("UPDATE plano SET status_plano='Inativo' WHERE data_fim < ?", (hoje,))
    conn.execute("UPDATE plano SET status_plano='Ativo'   WHERE data_fim >= ? AND status_plano != 'Suspenso'", (hoje,))
    # Atualiza status do aluno com base no plano
    conn.execute("""UPDATE aluno SET status='Ativo' WHERE ativo=1 AND id_aluno IN (
        SELECT id_aluno FROM plano WHERE data_fim >= ? AND status_plano='Ativo')""", (hoje,))
    conn.execute("""UPDATE aluno SET status='Inativo' WHERE ativo=1 AND id_aluno NOT IN (
        SELECT id_aluno FROM plano WHERE data_fim >= ? AND status_plano='Ativo')""", (hoje,))
    conn.commit()
    conn.close()


# ── Cadastro ─────────────────────────────────────────────────────
def cadastrar_aluno(dados: dict):
    cpf = dados["cpf"].replace(".", "").replace("-", "").strip()
    conn = get_connection()
    if conn.execute("SELECT 1 FROM aluno WHERE cpf=?", (cpf,)).fetchone():
        conn.close()
        return False, "CPF já cadastrado no sistema."
    try:
        cur = conn.execute(
            "INSERT INTO aluno (nome_completo,cpf,email,telefone,data_nascimento,genero,status) VALUES (?,?,?,?,?,?,?)",
            (dados["nome"], cpf, dados["email"], dados["telefone"],
             dados["nascimento"], dados["genero"], "Ativo")
        )
        id_aluno = cur.lastrowid
        tipo  = dados["plano"]
        hoje  = date.today()
        fim   = hoje + timedelta(days=DURACOES.get(tipo, 30))
        valor = VALORES.get(tipo, 0)
        conn.execute(
            "INSERT INTO plano (id_aluno,tipo_plano,data_inicio,data_fim,valor,status_plano) VALUES (?,?,?,?,?,?)",
            (id_aluno, tipo, hoje.isoformat(), fim.isoformat(), valor, "Ativo")
        )
        conn.commit()
        return True, f"Aluno '{dados['nome']}' cadastrado com sucesso!"
    except Exception as e:
        conn.rollback()
        return False, str(e)
    finally:
        conn.close()


# ── Listagem ─────────────────────────────────────────────────────
def listar_alunos(filtro=""):
    conn = get_connection()
    rows = conn.execute("""
        SELECT a.id_aluno, a.nome_completo, a.cpf, a.email, a.telefone,
               a.data_nascimento, a.genero, a.status,
               p.id_plano, p.tipo_plano, p.data_inicio, p.data_fim,
               p.valor, p.status_plano
        FROM aluno a
        LEFT JOIN (
            SELECT *, ROW_NUMBER() OVER (PARTITION BY id_aluno ORDER BY data_fim DESC) rn
            FROM plano
        ) p ON a.id_aluno=p.id_aluno AND p.rn=1
        WHERE a.ativo=1
        ORDER BY a.nome_completo
    """).fetchall()
    conn.close()
    result = [dict(r) for r in rows]
    if filtro:
        f = filtro.lower()
        result = [r for r in result if f in r["nome_completo"].lower() or f in (r["cpf"] or "")]
    return result


def get_aluno(id_aluno: int):
    conn = get_connection()
    row = conn.execute("""
        SELECT a.*, p.id_plano, p.tipo_plano, p.data_inicio, p.data_fim, p.valor, p.status_plano
        FROM aluno a
        LEFT JOIN (
            SELECT *, ROW_NUMBER() OVER (PARTITION BY id_aluno ORDER BY data_fim DESC) rn
            FROM plano
        ) p ON a.id_aluno=p.id_aluno AND p.rn=1
        WHERE a.id_aluno=?
    """, (id_aluno,)).fetchone()
    conn.close()
    return dict(row) if row else None


# ── Edição ───────────────────────────────────────────────────────
def editar_aluno(id_aluno: int, dados: dict):
    cpf = dados["cpf"].replace(".", "").replace("-", "").strip()
    conn = get_connection()
    try:
        dup = conn.execute("SELECT id_aluno FROM aluno WHERE cpf=? AND id_aluno!=?", (cpf, id_aluno)).fetchone()
        if dup:
            return False, "CPF já cadastrado para outro aluno."
        conn.execute(
            "UPDATE aluno SET nome_completo=?,cpf=?,email=?,telefone=?,data_nascimento=?,genero=? WHERE id_aluno=?",
            (dados["nome"], cpf, dados["email"], dados["telefone"],
             dados["nascimento"], dados["genero"], id_aluno)
        )
        tipo_novo = dados.get("plano")
        if tipo_novo:
            plano = conn.execute(
                "SELECT * FROM plano WHERE id_aluno=? ORDER BY data_fim DESC LIMIT 1", (id_aluno,)
            ).fetchone()
            if plano:
                if plano["tipo_plano"] != tipo_novo:
                    ini = date.fromisoformat(plano["data_inicio"])
                    fim = ini + timedelta(days=DURACOES.get(tipo_novo, 30))
                    conn.execute(
                        "UPDATE plano SET tipo_plano=?,data_fim=?,valor=? WHERE id_plano=?",
                        (tipo_novo, fim.isoformat(), VALORES.get(tipo_novo, 0), plano["id_plano"])
                    )
            else:
                hoje = date.today()
                fim  = hoje + timedelta(days=DURACOES.get(tipo_novo, 30))
                conn.execute(
                    "INSERT INTO plano (id_aluno,tipo_plano,data_inicio,data_fim,valor,status_plano) VALUES (?,?,?,?,?,?)",
                    (id_aluno, tipo_novo, hoje.isoformat(), fim.isoformat(), VALORES.get(tipo_novo, 0), "Ativo")
                )
        conn.commit()
        return True, "Aluno atualizado com sucesso!"
    except Exception as e:
        conn.rollback()
        return False, str(e)
    finally:
        conn.close()


# ── Toggle status do plano ───────────────────────────────────────
def alternar_status_plano(id_plano: int):
    """Alterna manualmente o status_plano entre Ativo e Suspenso."""
    conn = get_connection()
    try:
        plano = conn.execute("SELECT * FROM plano WHERE id_plano=?", (id_plano,)).fetchone()
        if not plano:
            return False, "Plano não encontrado."
        novo = "Suspenso" if plano["status_plano"] == "Ativo" else "Ativo"
        conn.execute("UPDATE plano SET status_plano=? WHERE id_plano=?", (novo, id_plano))
        # Recalcula status do aluno
        hoje = date.today().isoformat()
        id_aluno = plano["id_aluno"]
        tem_ativo = conn.execute(
            "SELECT 1 FROM plano WHERE id_aluno=? AND data_fim>=? AND status_plano='Ativo'",
            (id_aluno, hoje)
        ).fetchone()
        novo_status_aluno = "Ativo" if tem_ativo else "Inativo"
        conn.execute("UPDATE aluno SET status=? WHERE id_aluno=?", (novo_status_aluno, id_aluno))
        conn.commit()
        return True, novo
    finally:
        conn.close()


# ── Desativar (soft delete) ──────────────────────────────────────
def desativar_aluno(id_aluno: int):
    conn = get_connection()
    conn.execute("UPDATE aluno SET ativo=0, status='Inativo' WHERE id_aluno=?", (id_aluno,))
    conn.commit()
    conn.close()


# ── Métricas ─────────────────────────────────────────────────────
def metricas():
    hoje = date.today().isoformat()
    conn = get_connection()
    total  = conn.execute("SELECT COUNT(*) FROM aluno WHERE ativo=1").fetchone()[0]
    ativos = conn.execute(
        "SELECT COUNT(DISTINCT id_aluno) FROM plano WHERE data_fim>=? AND status_plano='Ativo'", (hoje,)
    ).fetchone()[0]
    pend   = conn.execute(
        "SELECT COUNT(*) FROM aluno WHERE ativo=1 AND status='Inativo'"
    ).fetchone()[0]
    conn.close()
    churn = round(pend / total * 100, 1) if total else 0
    return {"total_alunos": total, "planos_ativos": ativos,
            "renovacoes_pendentes": pend, "churn": churn}


def matriculas_por_mes():
    conn = get_connection()
    rows = conn.execute("""
        SELECT strftime('%m/%Y', data_inicio) AS mes,
               strftime('%Y%m', data_inicio)  AS ordem,
               COUNT(*) AS total
        FROM plano GROUP BY mes ORDER BY ordem
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows][-12:]


# ── Exportação ───────────────────────────────────────────────────
def exportar_csv():
    rows = listar_alunos()
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.rename(columns={
            "nome_completo": "Nome", "cpf": "CPF", "email": "E-mail",
            "telefone": "Telefone", "data_nascimento": "Nascimento",
            "genero": "Gênero", "status": "Status Aluno",
            "tipo_plano": "Plano", "data_inicio": "Início Plano",
            "data_fim": "Vencimento", "valor": "Valor (R$)",
            "status_plano": "Status Plano"
        })
        cols = ["Nome","CPF","E-mail","Telefone","Nascimento","Gênero",
                "Status Aluno","Plano","Início Plano","Vencimento","Valor (R$)","Status Plano"]
        df = df[[c for c in cols if c in df.columns]]
    path = REPORTS_DIR / f"relatorio_{date.today().isoformat()}.csv"
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return str(path)


def exportar_excel():
    rows = listar_alunos()
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.rename(columns={
            "nome_completo": "Nome", "cpf": "CPF", "email": "E-mail",
            "telefone": "Telefone", "data_nascimento": "Nascimento",
            "genero": "Gênero", "status": "Status Aluno",
            "tipo_plano": "Plano", "data_inicio": "Início Plano",
            "data_fim": "Vencimento", "valor": "Valor (R$)",
            "status_plano": "Status Plano"
        })
        cols = ["Nome","CPF","E-mail","Telefone","Nascimento","Gênero",
                "Status Aluno","Plano","Início Plano","Vencimento","Valor (R$)","Status Plano"]
        df = df[[c for c in cols if c in df.columns]]

    path = REPORTS_DIR / f"relatorio_{date.today().isoformat()}.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Alunos")
        ws = w.sheets["Alunos"]
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(c.value or "")) for c in col) + 4
    return str(path)
